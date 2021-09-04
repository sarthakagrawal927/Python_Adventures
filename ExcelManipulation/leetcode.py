from openpyxl import load_workbook, Workbook
workbook = load_workbook(filename="final.xlsx")

sheet = workbook.active

abc = 1
index = 2

newFile = "result.xlsx"

workbook2 = Workbook()
new_sheet = workbook2.active

cells = sheet['A1':'A15']

for n in range(1431):
    value = sheet.cell(row = n+1, column= 1).value

    if abc == 1:
        abc = 2
        index_name = value.split('.')
        new_sheet['B' + str(index)].hyperlink = sheet.cell(row = n+1, column= 1).hyperlink.target
        new_sheet['B' + str(index)].style = "Hyperlink"
        new_sheet['A' + str(index)] = index_name[0]
        new_sheet['B' + str(index)] = index_name[1]

    elif abc == 3:
        new_sheet["C" + str(index)] = str(value * 100)[:4] + "%"
    else:
        new_sheet["D" + str(index)] = value

    abc += 1
    if abc == 5:
        abc = 1
        index += 1

workbook2.save(filename=newFile)